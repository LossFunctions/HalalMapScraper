#!/usr/bin/env python3
"""
Instagram scraper for summarizing halal-focused posts.

Given one or more public Instagram accounts, the script fetches their
most recent posts through Instagram's web profile API, records the post
date, inferred restaurant, keywords mentioned, and optionally enforces the
halal + new-opening keyword filter before appending entries to a text
report ordered from newest to oldest post seen.
"""
from __future__ import annotations

import argparse
import csv
import datetime as _dt
import json
import os
import random
from collections import defaultdict
from email.utils import parsedate_to_datetime
from pathlib import Path
import re
import time
import warnings
from typing import Dict, Iterable, List, Optional, Sequence

warnings.filterwarnings(
    "ignore", message="urllib3 v2 only supports OpenSSL 1.1.1+", category=Warning
)

import requests


INSTAGRAM_WEB_PROFILE_URL = (
    "https://www.instagram.com/api/v1/users/web_profile_info/?username={username}"
)

REQUEST_TIMEOUT_SECONDS = 20
DEFAULT_LIMIT = 12
DEFAULT_MAX_RETRIES = 4
DEFAULT_MIN_REQUEST_INTERVAL = 1.2
DEFAULT_ACCOUNT_COOLDOWN_SECONDS = 1.5
DEFAULT_BACKOFF_BASE_SECONDS = 2.0
MAX_BACKOFF_SECONDS = 90.0
RETRYABLE_STATUS_CODES = {429, 500, 502, 503, 504}
DEFAULT_DEBUG_HTTP_LOG = Path("instagram_http_debug.jsonl")
DEBUG_RESPONSE_HEADERS = (
    "Retry-After",
    "Content-Type",
    "Content-Length",
    "X-FB-Request-ID",
    "X-FB-Trace-ID",
    "X-FB-Rev",
    "X-IG-Set-WWW-Claim",
    "X-IG-Origin-Region",
    "x-webcache-source",
    "Server",
    "CF-Ray",
)

# Default accounts file to read from when no accounts are supplied via CLI
DEFAULT_ACCOUNTS_FILE = Path("instagram_accounts.txt")

# Keywords: require "halal" plus at least one of the additional phrases.
DEFAULT_REQUIRED_KEYWORDS = ("halal",)
DEFAULT_NEWNESS_KEYWORDS = (
    "grand opening",
    "grand re-opening",
    "soft opening",
    "soft launch",
    "now open",
    "now serving halal",
    "just opened",
    "opened today",
    "opening weekend",
    "new location",
    "new halal spot",
    "new halal restaurant",
    "new halal menu",
    "latest halal",
    "brand new",
    "halal opening",
    "halal launch",
    "coming soon",
    "opening soon",
    "open now",
    "new arrivals",
)

NON_RESTAURANT_HANDLES = {
    "_rehansyed",
    "_nursenez_",
    "aaxotics",
    "adamsaleh",
    "adamthedunker",
    "albaydargroup",
    "centercitymosque",
    "champagneadan",
    "dmv3ats",
    "dropletsofmercyusa",
    "dropletsofmercy",
    "faiz.yy",
    "faizalfilli",
    "delicatehijabi",
    "halalfoodfestdmv",
    "islamicsocietyofbaltimore",
    "jerseyhalalspots",
    "majlisofny",
    "mdq_academy",
    "mdqyouth",
    "mymasjidal",
    "muslimlightsfest",
    "muscareinc",
    "mrs.fields__tcby",
    "movingimagesusinc",
    "omarthecarguy",
    "nychalalfest",
    "razadastgir",
    "y4.m33n",
    "thecarguysny",
    "phillyhalalfoodfest",
    "phillyhalalspots",
}


class FetchRequestError(RuntimeError):
    """Structured network failure details for troubleshooting."""

    def __init__(
        self,
        message: str,
        status_code: Optional[int] = None,
        endpoint_name: str = "",
        response_headers: Optional[Dict[str, str]] = None,
        response_preview: str = "",
    ) -> None:
        super().__init__(message)
        self.status_code = status_code
        self.endpoint_name = endpoint_name
        self.response_headers = response_headers or {}
        self.response_preview = response_preview


def utc_now_iso() -> str:
    return _dt.datetime.now(tz=_dt.timezone.utc).isoformat()


def compact_body_preview(text: str, max_chars: int = 260) -> str:
    compact = re.sub(r"\s+", " ", text or "").strip()
    if len(compact) > max_chars:
        return compact[:max_chars] + "..."
    return compact


def extract_response_debug_headers(response: requests.Response) -> Dict[str, str]:
    selected: Dict[str, str] = {}
    for key in DEBUG_RESPONSE_HEADERS:
        value = response.headers.get(key)
        if value:
            selected[key.lower()] = value
    return selected


def write_http_debug_event(
    enabled: bool,
    log_path: Optional[Path],
    event: Dict[str, object],
    debug_state: Optional[Dict[str, bool]] = None,
) -> None:
    if not enabled or log_path is None:
        return
    payload = {"ts_utc": utc_now_iso(), **event}
    try:
        if log_path.parent != Path("."):
            log_path.parent.mkdir(parents=True, exist_ok=True)
        with log_path.open("a", encoding="utf-8") as fh:
            fh.write(json.dumps(payload, ensure_ascii=False) + "\n")
    except OSError as exc:
        already_reported = bool(debug_state and debug_state.get("log_write_error_reported"))
        if already_reported:
            return
        if debug_state is not None:
            debug_state["log_write_error_reported"] = True
        print(f"[warn] Could not write HTTP debug log {log_path}: {exc}", flush=True)


def fetch_recent_posts(
    username: str,
    limit: int,
    session_headers: Optional[Dict[str, str]] = None,
    session_cookies: Optional[Dict[str, str]] = None,
    client: Optional[requests.Session] = None,
    max_retries: int = DEFAULT_MAX_RETRIES,
    min_request_interval: float = DEFAULT_MIN_REQUEST_INTERVAL,
    throttle_state: Optional[Dict[str, float]] = None,
    debug_http: bool = False,
    debug_http_log: Optional[Path] = None,
    debug_state: Optional[Dict[str, bool]] = None,
) -> List[dict]:
    """Return up to `limit` recent post nodes for a public username.

    Notes:
    - Instagram's web profile API returns only the first page (typically 12 posts).
      To avoid silently truncating results, we combine that page with the user
      feed endpoint which supports pagination to satisfy `limit`.
    """
    if limit <= 0:
        return []

    url = INSTAGRAM_WEB_PROFILE_URL.format(username=username)
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/123.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json",
        "Accept-Language": "en-US,en;q=0.9",
        "X-IG-App-ID": "936619743392459",
        "X-ASBD-ID": "129477",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": "https://www.instagram.com/",
    }

    if session_headers:
        headers.update(session_headers)

    max_retries = max(max_retries, 0)
    min_request_interval = max(min_request_interval, 0.0)
    cookies = session_cookies or {}
    session = client or requests.Session()
    local_throttle_state: Dict[str, float] = (
        throttle_state if throttle_state is not None else {}
    )

    payload = request_json_with_retries(
        session=session,
        url=url,
        headers=headers,
        cookies=cookies,
        username=username,
        endpoint_name="profile",
        max_retries=max_retries,
        min_request_interval=min_request_interval,
        throttle_state=local_throttle_state,
        debug_http=debug_http,
        debug_http_log=debug_http_log,
        debug_state=debug_state,
    )

    # Parse initial page edges and user id when available
    edges = []
    user_id = None
    try:
        user_obj = payload["data"]["user"]
        media_section = user_obj.get("edge_owner_to_timeline_media") or {}
        edges = media_section.get("edges", []) or []
        user_id = user_obj.get("id")
    except (KeyError, TypeError):
        pass

    nodes: List[dict] = []
    seen_shortcodes: set[str] = set()
    for edge in edges:
        try:
            node = edge["node"]
        except Exception:
            continue
        sc = (node.get("shortcode") or "").strip()
        if sc and sc not in seen_shortcodes:
            nodes.append(node)
            seen_shortcodes.add(sc)
        if len(nodes) >= limit:
            return nodes[:limit]

    # If we still need more posts and have a user_id, use the paginated feed
    if user_id:
        count = min(max(limit - len(nodes), 1), 50)
        feed_url = f"https://www.instagram.com/api/v1/feed/user/{user_id}/?count={count}"
        try:
            feed_payload = request_json_with_retries(
                session=session,
                url=feed_url,
                headers=headers,
                cookies=cookies,
                username=username,
                endpoint_name="feed page 1",
                max_retries=max_retries,
                min_request_interval=min_request_interval,
                throttle_state=local_throttle_state,
                debug_http=debug_http,
                debug_http_log=debug_http_log,
                debug_state=debug_state,
            )
        except RuntimeError as exc:
            print(
                f"[warn] {username}: {exc}. Returning {len(nodes)} profile posts.",
                flush=True,
            )
            return nodes[:limit]

        items = list(feed_payload.get("items", []) or [])
        next_max_id = feed_payload.get("next_max_id")
        more_available = feed_payload.get("more_available")

        while len(nodes) < limit and (items or (more_available and next_max_id)):
            for item in items:
                converted = convert_feed_item(item)
                sc = (converted.get("shortcode") or "").strip()
                if sc and sc not in seen_shortcodes:
                    nodes.append(converted)
                    seen_shortcodes.add(sc)
                    if len(nodes) >= limit:
                        break
            if len(nodes) >= limit or not (more_available and next_max_id):
                break
            remaining = min(max(limit - len(nodes), 1), 50)
            paged_url = (
                f"https://www.instagram.com/api/v1/feed/user/{user_id}/"
                f"?count={remaining}&max_id={next_max_id}"
            )
            try:
                paged_json = request_json_with_retries(
                    session=session,
                    url=paged_url,
                    headers=headers,
                    cookies=cookies,
                    username=username,
                    endpoint_name="feed page",
                    max_retries=max_retries,
                    min_request_interval=min_request_interval,
                    throttle_state=local_throttle_state,
                    debug_http=debug_http,
                    debug_http_log=debug_http_log,
                    debug_state=debug_state,
                )
            except RuntimeError as exc:
                print(
                    f"[warn] {username}: {exc}. Stopping pagination at {len(nodes)} posts.",
                    flush=True,
                )
                break
            items = list(paged_json.get("items", []) or [])
            next_max_id = paged_json.get("next_max_id")
            more_available = paged_json.get("more_available")

    return nodes[:limit]


def parse_retry_after_seconds(value: Optional[str]) -> Optional[float]:
    """Parse Retry-After header as seconds (integer or HTTP-date)."""
    if not value:
        return None

    candidate = value.strip()
    if not candidate:
        return None

    if candidate.isdigit():
        return max(float(candidate), 0.0)

    try:
        retry_at = parsedate_to_datetime(candidate)
        if retry_at.tzinfo is None:
            retry_at = retry_at.replace(tzinfo=_dt.timezone.utc)
        now = _dt.datetime.now(tz=_dt.timezone.utc)
        return max((retry_at - now).total_seconds(), 0.0)
    except (TypeError, ValueError, OverflowError):
        return None


def compute_backoff_seconds(
    response: Optional[requests.Response],
    attempt: int,
) -> float:
    """Compute wait time for retries using Retry-After when available."""
    retry_after = None
    if response is not None:
        retry_after = parse_retry_after_seconds(response.headers.get("Retry-After"))

    if retry_after is None:
        retry_after = min(
            DEFAULT_BACKOFF_BASE_SECONDS * (2**attempt),
            MAX_BACKOFF_SECONDS,
        )

    return min(retry_after + random.uniform(0.15, 0.75), MAX_BACKOFF_SECONDS)


def enforce_min_request_interval(
    min_request_interval: float,
    throttle_state: Dict[str, float],
) -> None:
    """Ensure requests are spaced out to reduce rate limit responses."""
    if min_request_interval <= 0:
        throttle_state["last_request_ts"] = time.monotonic()
        return

    now = time.monotonic()
    last_request_ts = throttle_state.get("last_request_ts")
    if last_request_ts is not None:
        elapsed = now - last_request_ts
        if elapsed < min_request_interval:
            sleep_seconds = (min_request_interval - elapsed) + random.uniform(0.05, 0.25)
            time.sleep(max(sleep_seconds, 0.0))
    throttle_state["last_request_ts"] = time.monotonic()


def request_json_with_retries(
    session: requests.Session,
    url: str,
    headers: Dict[str, str],
    cookies: Dict[str, str],
    username: str,
    endpoint_name: str,
    max_retries: int,
    min_request_interval: float,
    throttle_state: Dict[str, float],
    debug_http: bool = False,
    debug_http_log: Optional[Path] = None,
    debug_state: Optional[Dict[str, bool]] = None,
) -> dict:
    """GET JSON with retry/backoff for transient failures like 429/5xx."""
    last_error = "unknown error"
    last_status_code: Optional[int] = None
    last_headers: Dict[str, str] = {}
    last_body_preview = ""

    write_http_debug_event(
        enabled=debug_http,
        log_path=debug_http_log,
        debug_state=debug_state,
        event={
            "event": "request_batch_start",
            "username": username,
            "endpoint": endpoint_name,
            "url": url,
            "max_retries": max_retries,
            "min_request_interval": min_request_interval,
        },
    )

    for attempt in range(max_retries + 1):
        enforce_min_request_interval(min_request_interval, throttle_state)
        response: Optional[requests.Response] = None
        attempt_number = attempt + 1
        try:
            response = session.get(
                url,
                headers=headers,
                timeout=REQUEST_TIMEOUT_SECONDS,
                cookies=cookies,
            )
        except requests.RequestException as exc:
            last_error = f"{type(exc).__name__}: {exc}"
            write_http_debug_event(
                enabled=debug_http,
                log_path=debug_http_log,
                debug_state=debug_state,
                event={
                    "event": "request_exception",
                    "username": username,
                    "endpoint": endpoint_name,
                    "url": url,
                    "attempt": attempt_number,
                    "error": last_error,
                },
            )
        else:
            status = response.status_code
            retry_after_seconds = parse_retry_after_seconds(response.headers.get("Retry-After"))
            headers_subset = extract_response_debug_headers(response)
            body_preview = compact_body_preview(response.text or "")

            last_status_code = status
            last_headers = headers_subset
            last_body_preview = body_preview

            write_http_debug_event(
                enabled=debug_http,
                log_path=debug_http_log,
                debug_state=debug_state,
                event={
                    "event": "response",
                    "username": username,
                    "endpoint": endpoint_name,
                    "url": url,
                    "attempt": attempt_number,
                    "status_code": status,
                    "retry_after_seconds": retry_after_seconds,
                    "headers": headers_subset,
                    "body_preview": body_preview if status != requests.codes.ok else "",
                },
            )

            if response.status_code == requests.codes.ok:
                try:
                    return response.json()
                except json.JSONDecodeError as exc:
                    write_http_debug_event(
                        enabled=debug_http,
                        log_path=debug_http_log,
                        debug_state=debug_state,
                        event={
                            "event": "json_decode_error",
                            "username": username,
                            "endpoint": endpoint_name,
                            "url": url,
                            "attempt": attempt_number,
                            "status_code": status,
                            "error": str(exc),
                            "headers": headers_subset,
                            "body_preview": body_preview,
                        },
                    )
                    raise FetchRequestError(
                        f"Failed to parse JSON for '{username}' {endpoint_name}: {exc}. "
                        f"Body preview: {body_preview!r}",
                        status_code=status,
                        endpoint_name=endpoint_name,
                        response_headers=headers_subset,
                        response_preview=body_preview,
                    ) from exc

            last_error = f"HTTP {status}"
            retryable = status in RETRYABLE_STATUS_CODES
            if not retryable:
                raise FetchRequestError(
                    f"Failed to fetch {endpoint_name} for '{username}': HTTP {status}. "
                    f"Body preview: {body_preview!r}",
                    status_code=status,
                    endpoint_name=endpoint_name,
                    response_headers=headers_subset,
                    response_preview=body_preview,
                )

        if attempt >= max_retries:
            break

        wait_seconds = compute_backoff_seconds(response, attempt)
        print(
            f"[warn] {username}: {endpoint_name} {last_error}. "
            f"Retrying in {wait_seconds:.1f}s ({attempt + 1}/{max_retries}).",
            flush=True,
        )
        write_http_debug_event(
            enabled=debug_http,
            log_path=debug_http_log,
            debug_state=debug_state,
            event={
                "event": "retry_scheduled",
                "username": username,
                "endpoint": endpoint_name,
                "url": url,
                "attempt": attempt_number,
                "wait_seconds": round(wait_seconds, 3),
                "last_error": last_error,
                "status_code": last_status_code,
                "headers": last_headers,
            },
        )
        time.sleep(wait_seconds)

    raise FetchRequestError(
        f"Failed to fetch {endpoint_name} for '{username}' after {max_retries + 1} "
        f"attempts ({last_error}).",
        status_code=last_status_code,
        endpoint_name=endpoint_name,
        response_headers=last_headers,
        response_preview=last_body_preview,
    )


def convert_feed_item(item: dict) -> dict:
    """Normalize feed/user items to the graph-style node structure."""
    caption_obj = item.get("caption") or {}
    caption_text = ""
    if isinstance(caption_obj, dict):
        caption_text = caption_obj.get("text") or ""

    tag_edges = []
    usertags = item.get("usertags", {})
    if isinstance(usertags, dict):
        for tag in usertags.get("in", []) or []:
            username = (
                tag.get("user", {})
                if isinstance(tag, dict)
                else {}
            )
            if isinstance(username, dict):
                uname = username.get("username")
            else:
                uname = None
            if uname:
                tag_edges.append({"node": {"user": {"username": uname}}})

    location = item.get("location") or {}

    node = {
        "shortcode": item.get("code") or "",
        "taken_at_timestamp": item.get("taken_at"),
        "edge_media_to_caption": (
            {"edges": [{"node": {"text": caption_text}}]}
            if caption_text
            else {"edges": []}
        ),
        "edge_media_to_tagged_user": {"edges": tag_edges},
        "location": location,
    }
    return node


def extract_caption(node: dict) -> str:
    """Pull the caption text from a post node."""
    edges = node.get("edge_media_to_caption", {}).get("edges", [])
    if not edges:
        return ""
    return edges[0]["node"].get("text", "") or ""


def pick_place_name(node: dict, account_username: str, caption: str) -> Optional[str]:
    """Try to infer the place name from location metadata, tags, or caption."""
    location = node.get("location") or {}
    location_name = location.get("name")
    if isinstance(location_name, str) and location_name.strip():
        return location_name.strip()

    tagged_users = node.get("edge_media_to_tagged_user", {}).get("edges", [])
    for tag in tagged_users:
        username = (
            tag.get("node", {})
            .get("user", {})
            .get("username", "")
            .strip()
        )
        if username and username.lower() != account_username.lower():
            return f"@{username}"

    caption_line = next((line.strip() for line in caption.splitlines() if line.strip()), "")
    return caption_line or None


def _keyword_in_text(keyword: str, text: str) -> bool:
    """Check if `keyword` exists in `text`, respecting word boundaries for single words."""
    keyword_lc = keyword.lower()
    text_lc = text.lower()
    if " " in keyword_lc:
        return keyword_lc in text_lc
    pattern = rf"\\b{re.escape(keyword_lc)}\\b"
    return re.search(pattern, text_lc) is not None


FULL_ADDRESS_REGEX = re.compile(
    r"\d{1,5}[-\w\s'&\.]*?(?:Street|St|Avenue|Ave|Road|Rd|Boulevard|Blvd|Lane|Ln|Drive|Dr|Court|Ct|Place|Pl|Way|Parkway|Pkwy|Highway|Hwy|Terrace|Ter|Trail|Trl|Circle|Cir|Center|Ctr)"
    r"(?:[-\w\s'&\.]*)?,\s*[A-Za-z .'-]+,\s*[A-Z]{2}(?:\s*\d{5})?",
    re.IGNORECASE,
)
STREET_FALLBACK_REGEX = re.compile(
    r"\d{1,5}[-\w\s'&\.]*(?:Street|St|Avenue|Ave|Road|Rd|Boulevard|Blvd|Lane|Ln|Drive|Dr|Court|Ct|Place|Pl|Way|Parkway|Pkwy|Highway|Hwy|Terrace|Ter|Trail|Trl|Circle|Cir|Center|Ctr)\b",
    re.IGNORECASE,
)
CITY_STATE_REGEX = re.compile(
    r"\b(?:New York|NYC|Brooklyn|Queens|Bronx|Staten Island|Manhattan|Long Island|Jersey City|New Jersey|NJ|Philadelphia|PA|Connecticut|CT|Boston|MA|Maryland|MD|Virginia|VA)\b",
    re.IGNORECASE,
)
VENUE_PREFIX_CLEAN = re.compile(
    r"^(?:located|location|address|find us(?: at)?|pull up to|come through to|come thru to|come to|stop by|find them at|meet us at|visit us(?: at)?|come through)\s*[:\-]*\s*",
    re.IGNORECASE,
)
VENUE_SKIP_WORDS = re.compile(
    r"\b(?:grand opening|soft opening|today|tonight|from|until|till|oct|nov|dec|jan|feb|mar|apr|may|jun|jul|aug|sep|sept|pm|am)\b",
    re.IGNORECASE,
)
TIME_TEXT_PATTERN = re.compile(
    r"\b(?:[0-9]{1,2}\s*(?:am|pm)|[0-9]{1,2}[:.][0-9]{2}\s*(?:am|pm)|today|tonight|from\s+[0-9]|until\s+[0-9]|till\s+[0-9])\b",
    re.IGNORECASE,
)
SENTENCE_BREAK_TOKENS = [
    " #",
    " @",
    " http",
    " www",
    " Follow",
    " Give ",
    " Items",
    " MENU",
    " Menu",
    " DM ",
    " Enjoy",
    " Call",
    " Text",
    " Order",
    " Visit",
    " Check",
]


def infer_location_from_caption(
    caption: str,
) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """Attempt to extract venue/address/city hints from caption text."""
    venue_candidate: Optional[str] = None
    address_candidate: Optional[str] = None
    city_candidate: Optional[str] = None
    address_confidence = -1  # -1 none, 0 weak, 1 fallback, 2 full

    def _maybe_set_city(address_text: str) -> None:
        nonlocal city_candidate
        if city_candidate:
            return
        parts = [part.strip() for part in address_text.split(",") if part.strip()]
        if len(parts) >= 2:
            city_candidate = ", ".join(parts[1:])

    for raw_line in caption.splitlines():
        line = raw_line.strip(" -*•")
        if not line:
            continue
        line = line.lstrip("📍").strip()
        if not line:
            continue
        line = VENUE_PREFIX_CLEAN.sub("", line)
        line_clean = re.sub(r"\s+", " ", line)

        def consider_venue(text: str) -> None:
            nonlocal venue_candidate
            candidate = re.sub(r"\(@[^)]+\)", "", text).strip(" ,:-")
            if not candidate:
                return
            if VENUE_SKIP_WORDS.search(candidate):
                return
            if not venue_candidate:
                venue_candidate = candidate

        if not any(char.isdigit() for char in line_clean):
            if "@" in line_clean:
                consider_venue(line_clean)
            continue

        address_text = None
        prefix_text = None
        confidence = 0
        match = FULL_ADDRESS_REGEX.search(line_clean)
        if match:
            address_text = match.group(0).strip(" ,")
            prefix_text = line_clean[: match.start()].strip(" ,:-")
            confidence = 2
        else:
            fallback_match = STREET_FALLBACK_REGEX.search(line_clean)
            if fallback_match:
                start = fallback_match.start()
                end = fallback_match.end()
                prefix_text = line_clean[:start].strip(" ,:-")
                substring = line_clean[start:]
                confidence = 1
            else:
                digit_index = next((i for i, ch in enumerate(line_clean) if ch.isdigit()), -1)
                if digit_index == -1:
                    continue
                prefix_text = line_clean[:digit_index].strip(" ,:-")
                substring = line_clean[digit_index:]
                confidence = 0

            substring = substring.strip()
            for token in SENTENCE_BREAK_TOKENS:
                pos = substring.find(token)
                if pos > 5:
                    substring = substring[:pos]
                    break
            if ". " in substring:
                substring = substring.split(". ", 1)[0]
            address_text = substring.strip(" ,")
            if TIME_TEXT_PATTERN.search(address_text) and confidence < 1:
                address_text = None

        if address_text and confidence > address_confidence:
            address_candidate = address_text
            address_confidence = confidence
            _maybe_set_city(address_text)

        if prefix_text:
            consider_venue(prefix_text)

        if not city_candidate:
            city_match = CITY_STATE_REGEX.search(line_clean)
            if city_match:
                city_candidate = city_match.group().strip(" ,")

    return venue_candidate, address_candidate, city_candidate


def process_posts(
    username: str,
    nodes: Sequence[dict],
    required_keywords: Iterable[str],
    newness_keywords: Iterable[str],
    require_keywords: bool,
) -> List[dict]:
    """Extract metadata for post nodes, optionally filtering by keyword rules."""
    required = tuple(k.lower() for k in required_keywords)
    newness = tuple(k.lower() for k in newness_keywords)

    processed: List[dict] = []
    for node in nodes:
        caption = extract_caption(node)
        caption_lower = caption.lower()

        if require_keywords:
            if not caption_lower:
                continue
            if not all(_keyword_in_text(keyword, caption_lower) for keyword in required):
                continue
            if not any(_keyword_in_text(keyword, caption_lower) for keyword in newness):
                continue

        matched_keywords = sorted(
            {
                kw
                for kw in (*required, *newness)
                if caption_lower and _keyword_in_text(kw, caption_lower)
            }
        )

        post_url = f"https://www.instagram.com/p/{node.get('shortcode', '').strip()}/"
        timestamp = node.get("taken_at_timestamp")
        post_datetime = (
            _dt.datetime.fromtimestamp(timestamp)
            if isinstance(timestamp, (int, float))
            else None
        )

        post_date = (
            post_datetime.date().isoformat()
            if isinstance(post_datetime, _dt.datetime)
            else None
        )
        location_obj = node.get("location") or {}
        location_name = location_obj.get("name")
        location_address = None
        location_city = location_obj.get("city")
        location_lat = location_obj.get("lat")
        location_lng = location_obj.get("lng")
        address_json = location_obj.get("address_json")
        if isinstance(address_json, str) and address_json:
            try:
                address_parsed = json.loads(address_json)
            except json.JSONDecodeError:
                address_parsed = {}
            if isinstance(address_parsed, dict):
                location_address = (
                    address_parsed.get("street_address")
                    or address_parsed.get("address_line1")
                )
                location_city = location_city or address_parsed.get("city_name")

        caption_venue, inferred_address, inferred_city = infer_location_from_caption(caption)
        if not location_address and inferred_address:
            location_address = inferred_address
        if not location_city and inferred_city:
            location_city = inferred_city

        tagged_edges = node.get("edge_media_to_tagged_user", {}).get("edges", [])
        tagged_accounts = []
        for tag in tagged_edges:
            username_tag = (
                tag.get("node", {})
                .get("user", {})
                .get("username", "")
            )
            if (
                username_tag
                and username_tag.lower() != username.lower()
                and username_tag not in tagged_accounts
            ):
                tagged_accounts.append(username_tag)

        place_name = pick_place_name(node, username, caption)
        caption_venue_clean = (caption_venue or "").strip()
        if caption_venue_clean:
            if (
                not place_name
                or place_name.startswith("@")
                or place_name.lower() == username.lower()
                or place_name.lower().startswith("new ")
                or place_name.lower().startswith("grand opening")
            ):
                place_name = caption_venue_clean

        processed.append(
            {
                "account": username,
                "post_url": post_url,
                "caption": caption.strip(),
                "place": place_name,
                "timestamp": timestamp,
                "datetime": post_datetime.isoformat() if post_datetime else None,
                "date": post_date,
                "keywords": matched_keywords,
                "location_name": location_name,
                "location_address": location_address,
                "location_city": location_city,
                "location_lat": location_lat,
                "location_lng": location_lng,
                "tagged_accounts": tagged_accounts,
                "caption_venue": caption_venue_clean or None,
            }
        )

    # Keep newest posts first based on timestamp if present.
    return sorted(processed, key=lambda item: item.get("timestamp") or 0, reverse=True)


def _clean_caption(text: str) -> str:
    """Collapse whitespace to keep captions readable in single-line outputs."""
    if not text:
        return ""
    return re.sub(r"\s+", " ", text).strip()


def load_existing_records(path: Path) -> List[dict]:
    if not path.exists() or path.stat().st_size == 0:
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, list):
            return data
    except (json.JSONDecodeError, OSError):
        pass
    return []


def merge_records(existing: Sequence[dict], new_records: Sequence[dict]) -> List[dict]:
    merged: dict[str, dict] = {}
    order: List[str] = []

    def upsert(record: dict) -> None:
        key = record.get("post_url") or f"{record.get('account')}|{record.get('timestamp')}"
        if key in merged:
            merged[key] = record
        else:
            merged[key] = record
            order.append(key)

    for item in existing:
        upsert(item)
    for item in new_records:
        upsert(item)

    ordered_records = [merged[key] for key in order]
    ordered_records.sort(key=lambda item: item.get("timestamp") or 0, reverse=True)
    return ordered_records


def write_json(records: Sequence[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as handle:
        json.dump(records, handle, indent=2, ensure_ascii=False)


def write_csv(records: Sequence[dict], output_path: Path) -> None:
    """Write post details to CSV for spreadsheet analysis."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "captured_date",
        "posted_date",
        "account",
        "keywords",
        "venue",
        "post_url",
        "location_address",
        "location_city",
        "tagged_accounts",
        "caption",
    ]

    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for item in records:
            keywords = "; ".join(item.get("keywords") or [])
            venue = item.get("place") or item.get("location_name") or ""
            row = {
                "captured_date": item.get("run_date") or "",
                "posted_date": item.get("date") or "",
                "account": item.get("account") or "",
                "keywords": keywords,
                "venue": venue,
                "post_url": item.get("post_url") or "",
                "location_address": item.get("location_address") or "",
                "location_city": item.get("location_city") or "",
                "tagged_accounts": "; ".join(item.get("tagged_accounts") or []),
                "caption": _clean_caption(item.get("caption", "")),
            }
            writer.writerow(row)


def write_excel(
    records: Sequence[dict],
    output_path: Path,
    monitored_accounts: Optional[Sequence[str]] = None,
) -> None:
    """Write an Excel workbook with highlighted keyword hits."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
    except ImportError:
        print("[warn] openpyxl not installed; skipping Excel export.")
        return

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Halal Openings"

    headers = [
        "captured_date",
        "posted_date",
        "account",
        "keywords",
        "venue",
        "post_url",
        "location_address",
        "location_city",
        "tagged_accounts",
        "caption",
    ]
    ws.append(headers)

    highlight = PatternFill(start_color="C7F5B4", end_color="C7F5B4", fill_type="solid")

    for item in records:
        keywords = "; ".join(item.get("keywords") or [])
        row = [
            item.get("run_date") or "",
            item.get("date") or "",
            item.get("account") or "",
            keywords,
            item.get("place") or item.get("location_name") or "",
            item.get("post_url") or "",
            item.get("location_address") or "",
            item.get("location_city") or "",
            "; ".join(item.get("tagged_accounts") or []),
            _clean_caption(item.get("caption", "")),
        ]
        ws.append(row)
        if keywords:
            for cell in ws[ws.max_row]:
                cell.fill = highlight

    accounts_to_ignore = {acc.lower() for acc in (monitored_accounts or [])} | NON_RESTAURANT_HANDLES
    accounts_sheet = wb.create_sheet("Tagged Accounts")
    accounts_sheet.append(
        ["tagged_account", "times_tagged", "latest_posted_date", "accounts_posted", "post_urls"]
    )

    tagged_map: Dict[str, Dict[str, set]] = defaultdict(
        lambda: {"posts": set(), "accounts": set(), "dates": set()}
    )
    for record in records:
        account = record.get("account") or ""
        post_url = record.get("post_url") or ""
        posted_date = record.get("date") or ""
        for tag in record.get("tagged_accounts") or []:
            if tag.lower() in accounts_to_ignore:
                continue
            key = tag.strip()
            if not key:
                continue
            tagged_map[key]["posts"].add(post_url)
            tagged_map[key]["accounts"].add(account)
            if posted_date:
                tagged_map[key]["dates"].add(posted_date)

    for tag, info in sorted(tagged_map.items()):
        latest_date = ""
        if info["dates"]:
            latest_date = max(info["dates"])
        accounts_list = "; ".join(sorted(acc for acc in info["accounts"] if acc))
        post_list = "; ".join(sorted(url for url in info["posts"] if url))
        accounts_sheet.append(
            [
                tag,
                len([url for url in info["posts"] if url]),
                latest_date,
                accounts_list,
                post_list,
            ]
        )

    wb.save(output_path)


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Summarize recent Instagram posts with optional halal keyword filtering."
    )
    parser.add_argument(
        "--accounts",
        nargs="+",
        help="One or more Instagram usernames to scan (e.g. halaleatsig).",
    )
    parser.add_argument(
        "--accounts-file",
        type=Path,
        help=(
            "Optional path to a text file with one username per line. "
            "If omitted, the script will use 'instagram_accounts.txt' when present."
        ),
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=DEFAULT_LIMIT,
        help=f"Number of recent posts per account to inspect (default: {DEFAULT_LIMIT}).",
    )
    parser.add_argument(
        "--max-retries",
        type=int,
        default=DEFAULT_MAX_RETRIES,
        help=(
            "Retries per request for transient failures like HTTP 429/5xx "
            f"(default: {DEFAULT_MAX_RETRIES})."
        ),
    )
    parser.add_argument(
        "--min-request-interval",
        type=float,
        default=DEFAULT_MIN_REQUEST_INTERVAL,
        help=(
            "Minimum delay in seconds between Instagram HTTP requests "
            f"(default: {DEFAULT_MIN_REQUEST_INTERVAL})."
        ),
    )
    parser.add_argument(
        "--account-cooldown",
        type=float,
        default=DEFAULT_ACCOUNT_COOLDOWN_SECONDS,
        help=(
            "Delay in seconds between accounts to reduce rate limits "
            f"(default: {DEFAULT_ACCOUNT_COOLDOWN_SECONDS})."
        ),
    )
    parser.add_argument(
        "--debug-http",
        action="store_true",
        help=(
            "Write HTTP troubleshooting events as JSONL "
            f"(default log: {DEFAULT_DEBUG_HTTP_LOG})."
        ),
    )
    parser.add_argument(
        "--debug-http-log",
        type=Path,
        default=DEFAULT_DEBUG_HTTP_LOG,
        help=f"Path for --debug-http JSONL output (default: {DEFAULT_DEBUG_HTTP_LOG}).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("halal_openings.json"),
        help="JSON file to store aggregated matches (default: halal_openings.json).",
    )
    parser.add_argument(
        "--csv-output",
        type=Path,
        default=Path("halal_openings.csv"),
        help="CSV file to append matches to (default: halal_openings.csv).",
    )
    parser.add_argument(
        "--xlsx-output",
        type=Path,
        default=Path("halal_openings.xlsx"),
        help="Optional Excel workbook to overwrite with the latest results.",
    )
    parser.add_argument(
        "--sessionid",
        help="Instagram sessionid cookie value (or set IG_SESSIONID env var).",
    )
    parser.add_argument(
        "--csrftoken",
        help="Instagram csrftoken cookie value (or set IG_CSRFTOKEN env var).",
    )
    parser.add_argument(
        "--require-keywords",
        action="store_true",
        help="Only record posts that include halal + new-opening keywords.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print matches to stdout without writing the output file.",
    )
    parser.add_argument(
        "--show-json",
        action="store_true",
        help="Dump match JSON to stdout for debugging.",
    )
    return parser.parse_args(argv)


def load_accounts(args: argparse.Namespace) -> List[str]:
    """Load usernames from CLI and/or a default file; normalize @prefixes."""
    accounts: List[str] = []

    # Direct CLI usernames
    if args.accounts:
        accounts.extend(args.accounts)

    # Prefer explicit file if provided; otherwise fall back to DEFAULT_ACCOUNTS_FILE if present
    files_to_read: List[Path] = []
    if args.accounts_file:
        files_to_read.append(args.accounts_file)
    elif not args.accounts and DEFAULT_ACCOUNTS_FILE.exists():
        files_to_read.append(DEFAULT_ACCOUNTS_FILE)

    for file_path in files_to_read:
        try:
            for line in file_path.read_text(encoding="utf-8").splitlines():
                raw = line.strip()
                # Skip comments and blanks
                if not raw or raw.startswith("#"):
                    continue
                # Strip leading @ if present
                if raw.startswith("@"):  # tolerate @username format
                    raw = raw[1:]
                accounts.append(raw)
        except OSError:
            # Ignore unreadable/missing files silently; caller handles empty case
            pass

    # Normalize and dedupe
    return sorted({acc.lstrip("@").lower() for acc in accounts if acc})


def normalize_cookie_value(raw_value: Optional[str], cookie_name: str) -> str:
    """Normalize cookie input from plain value or 'name=value; ...' string."""
    if not raw_value:
        return ""

    value = str(raw_value).strip()
    if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
        value = value[1:-1].strip()

    target = cookie_name.strip().lower()
    if not target:
        return value

    # Support pasting full cookie headers or single key/value strings.
    if ";" in value or "=" in value:
        for part in value.split(";"):
            piece = part.strip()
            if not piece or "=" not in piece:
                continue
            key, val = piece.split("=", 1)
            if key.strip().lower() == target:
                value = val.strip()
                break
        else:
            prefix = f"{target}="
            if value.lower().startswith(prefix):
                value = value[len(prefix) :].strip()

    value = value.strip().strip(";").strip()
    if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
        value = value[1:-1].strip()

    return value


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    accounts = load_accounts(args)
    if not accounts:
        print("No accounts provided. Use --accounts or --accounts-file.", flush=True)
        return 1

    raw_sessionid = args.sessionid or os.getenv("IG_SESSIONID")
    raw_csrftoken = args.csrftoken or os.getenv("IG_CSRFTOKEN")
    sessionid = normalize_cookie_value(raw_sessionid, "sessionid")
    csrftoken = normalize_cookie_value(raw_csrftoken, "csrftoken")
    session_cookies: Dict[str, str] = {}
    session_headers: Dict[str, str] = {}

    if sessionid:
        session_cookies["sessionid"] = sessionid
    if csrftoken:
        session_cookies["csrftoken"] = csrftoken
        session_headers["X-CSRFToken"] = csrftoken

    if session_cookies:
        session_headers.setdefault("X-Requested-With", "XMLHttpRequest")

    max_retries = max(args.max_retries, 0)
    min_request_interval = max(args.min_request_interval, 0.0)
    account_cooldown = max(args.account_cooldown, 0.0)
    debug_http = bool(args.debug_http)
    debug_http_log = args.debug_http_log
    http_session = requests.Session()
    throttle_state: Dict[str, float] = {}
    debug_state: Dict[str, bool] = {}

    if debug_http:
        print(
            f"[debug] HTTP diagnostics enabled. Writing JSONL events to {debug_http_log}.",
            flush=True,
        )
        write_http_debug_event(
            enabled=True,
            log_path=debug_http_log,
            debug_state=debug_state,
            event={
                "event": "run_context",
                "accounts_count": len(accounts),
                "limit": args.limit,
                "max_retries": max_retries,
                "min_request_interval": min_request_interval,
                "account_cooldown": account_cooldown,
                "has_sessionid": bool(sessionid),
                "has_csrftoken": bool(csrftoken),
                "sessionid_len": len(sessionid),
                "csrftoken_len": len(csrftoken),
                "sessionid_was_normalized": bool(raw_sessionid and str(raw_sessionid).strip() != sessionid),
                "csrftoken_was_normalized": bool(raw_csrftoken and str(raw_csrftoken).strip() != csrftoken),
            },
        )
        print(
            "[debug] Auth cookie summary: "
            f"sessionid={'yes' if bool(sessionid) else 'no'} (len={len(sessionid)}), "
            f"csrftoken={'yes' if bool(csrftoken) else 'no'} (len={len(csrftoken)}).",
            flush=True,
        )

    run_date = _dt.date.today().isoformat()
    all_records: List[dict] = []
    rate_limited_accounts = 0
    consecutive_429 = 0
    for idx, account in enumerate(accounts):
        try:
            nodes = fetch_recent_posts(
                account,
                args.limit,
                session_headers=session_headers or None,
                session_cookies=session_cookies or None,
                client=http_session,
                max_retries=max_retries,
                min_request_interval=min_request_interval,
                throttle_state=throttle_state,
                debug_http=debug_http,
                debug_http_log=debug_http_log,
                debug_state=debug_state,
            )
            consecutive_429 = 0
        except Exception as exc:
            print(f"[error] {account}: {exc}", flush=True)
            status_code = getattr(exc, "status_code", None)
            if status_code == 429:
                rate_limited_accounts += 1
                consecutive_429 += 1
                retry_after = ""
                origin_region = ""
                content_type = ""
                body_preview = ""
                if isinstance(exc, FetchRequestError):
                    retry_after = exc.response_headers.get("retry-after", "")
                    origin_region = exc.response_headers.get("x-ig-origin-region", "")
                    content_type = exc.response_headers.get("content-type", "")
                    body_preview = exc.response_preview
                print(
                    f"[diag] {account}: 429 details "
                    f"(retry-after={retry_after or 'n/a'}, "
                    f"origin-region={origin_region or 'n/a'}, "
                    f"content-type={content_type or 'n/a'}).",
                    flush=True,
                )
                if body_preview:
                    print(f"[diag] {account}: body preview: {body_preview!r}", flush=True)
                if consecutive_429 == 3:
                    print(
                        "[hint] 3 accounts in a row are rate-limited (HTTP 429). "
                        "This usually indicates temporary IP/session throttling by Instagram.",
                        flush=True,
                    )
                    print(
                        "[hint] Try waiting 30-60 minutes, increasing delays, lowering post limit, "
                        "or switching network/IP.",
                        flush=True,
                    )
            else:
                consecutive_429 = 0
            if idx < len(accounts) - 1 and account_cooldown > 0:
                time.sleep(account_cooldown + random.uniform(0.0, 0.35))
            continue

        records = process_posts(
            username=account,
            nodes=nodes,
            required_keywords=DEFAULT_REQUIRED_KEYWORDS,
            newness_keywords=DEFAULT_NEWNESS_KEYWORDS,
            require_keywords=args.require_keywords,
        )
        for record in records:
            record["run_date"] = run_date

        if not records:
            if args.require_keywords:
                print(f"[info] {account}: no posts matched the keyword rules.", flush=True)
            else:
                print(f"[info] {account}: no posts were collected.", flush=True)
            if idx < len(accounts) - 1 and account_cooldown > 0:
                time.sleep(account_cooldown + random.uniform(0.0, 0.35))
            continue

        keyword_hits = sum(1 for item in records if item.get("keywords"))
        print(
            f"[info] {account}: recorded {len(records)} posts "
            f"(keyword hits: {keyword_hits}).",
            flush=True,
        )
        all_records.extend(records)

        if args.show_json:
            print(json.dumps(records, indent=2, ensure_ascii=False))

        if idx < len(accounts) - 1 and account_cooldown > 0:
            time.sleep(account_cooldown + random.uniform(0.0, 0.35))

    if rate_limited_accounts:
        print(
            f"[info] Accounts rate-limited this run: {rate_limited_accounts}/{len(accounts)}.",
            flush=True,
        )
        if debug_http:
            print(
                f"[info] Review HTTP diagnostics in {debug_http_log} for response headers/body previews.",
                flush=True,
            )

    if args.dry_run:
        if all_records:
            print(json.dumps(all_records, indent=2, ensure_ascii=False))
        else:
            message = (
                "No posts matched the keyword criteria."
                if args.require_keywords
                else "No posts recorded."
            )
            print(message)
        return 0

    if not all_records:
        if args.require_keywords:
            print("No posts matched the keyword criteria.")
        else:
            print("No posts recorded.")
        return 0

    existing_records = load_existing_records(args.output)
    merged_records = merge_records(existing_records, all_records)

    write_json(merged_records, args.output)
    write_csv(merged_records, args.csv_output)
    if args.xlsx_output:
        write_excel(merged_records, args.xlsx_output, monitored_accounts=accounts)

    matched_accounts = sorted({item["account"] for item in merged_records})
    new_accounts = ", ".join(sorted({item["account"] for item in all_records}) or ["none"])
    print(
        f"Merged {len(merged_records)} total entries across accounts: {', '.join(matched_accounts)} "
        f"(added {len(all_records)} new from: {new_accounts}) "
        f"into {args.output}, {args.csv_output}"
        + (f", and {args.xlsx_output}" if args.xlsx_output else ""),
        flush=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
